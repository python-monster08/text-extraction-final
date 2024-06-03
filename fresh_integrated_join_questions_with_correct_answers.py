# this joins the extracted questions with the extracted correct answers
# and creates a dictionary called f'{base_filename_questions_modified}_questions_with_correct_answers_dictionary
import pandas as pd
import os
import importlib.util
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import importlib.util
from openpyxl.worksheet.datavalidation import DataValidation
import xlsxwriter
import re
# Folder path where the dictionary files are located
# folder_path = Path(r'C:\Users\PC\PycharmProjects\pythonProject')

# Function to load a module from a file path
# def load_module(file_name):
def load_module(file_name, folder_path):
    file_path = folder_path / file_name
    spec = importlib.util.spec_from_file_location("module.name", str(file_path))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module



def load_dictionary(file_name, base_dir_path):
    """ Load a dictionary from a Python file within the specified directory. """
    dict_path = base_dir_path / f"{file_name}.py"
    if not dict_path.exists():
        print(f"File {dict_path} not found.")
        return {}  # Return an empty dictionary instead of None

    spec = importlib.util.spec_from_file_location(file_name, dict_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    # Ensure the dictionary name does not include the '.py' extension
    dict_name = file_name if not file_name.endswith('.py') else file_name[:-3]

    # Attempt to get the dictionary
    loaded_dict = getattr(module, dict_name, {})

    # Check if the dictionary is empty and print a message
    if not loaded_dict:
        print(f"Loaded dictionary named '{dict_name}' is empty.")
    else:
        print(f"Loaded dictionary named '{dict_name}' successfully.")

    return loaded_dict


def extract_year_and_number(question):
    # Extract the year using a regular expression
    # year_match = re.search(r'\[(\d{4})\]', question)
    # Extract the year using a regular expression to handle both [2014] and [2014 - I] formats
    year_match = re.search(r'\[(\d{4})(?: - I)?\]', question)
    year = year_match.group(1) if year_match else '0000'
    if year_match:
        year = year_match.group(1)
        # Check if the year is not between 1947 and the current year
        if not (1947 <= int(year) <= datetime.now().year):
            year = '0000'
    else:
        year = '0000'  # Default to '0000' if no year is found

    # Extract the question number, assuming it's at the start and followed by a period or space
    number_match = re.search(r'^\s*(\d+)\.', question)

    if number_match:
        question_number = number_match.group(1)
        question_number = int(question_number)
        # print(f"The Question Number is {question_number}")
        # print(f"The integer extracted from Question Number is {question_number}")
        # user_input = input("Should I proceed with processing the Question Numbers? (y/n) ")
        # if user_input.lower() == 'y':
        # Check if the question number is not between 0 and 1000
        if not (0 <= question_number <= 1000):
                question_number = 9999
        # else:
        #    print("Not proceeding further.")
    else:
        question_number = 'AAA'  # Default to 'XXX' if no valid number is found

    return year, question_number


# Function to generate unique code for each question
def generate_unique_code(exam_name, exam_stage, subject_name, area_name, year, question_number):
    return f"{exam_name}_{exam_stage}_{subject_name}_{area_name}_{year}_{question_number}"


def create_excel_for_dict(data_dict, output_dir, filename_prefix):
    # Convert dictionary to DataFrame
    # df = pd.DataFrame(data_dict.values())
    # Convert dictionary to DataFrame; ensure 'question_id' is included by using items() and constructing a DataFrame directly
    df = pd.DataFrame([{'question_id': k, **v} for k, v in data_dict.items()])

    # Specify the file path to match the dictionary name
    # file_path = output_dir / f"{filename_prefix}.xlsx"
    # to also handle Hindi data
    file_path = output_dir / f"{filename_prefix}.xlsx"

    # Inform the user about the Excel file creation details before creating it
    print(f"Excel file will be created with the name '{filename_prefix}.xlsx' in the directory '{output_dir}'.")
    # input(f"Press Enter to Continue to create the Excel file for the dictionary with Excel file named: {filename_prefix}")

    # Write DataFrame to Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    # Confirm the creation of the file
    print(f"Excel file has been created and saved in '{file_path}'.")

def join_questions_with_answers(folder_path, questions_file, answers_file):

    questions_file_path = folder_path / questions_file
    answers_file_path = folder_path / answers_file

    # Check if the files exist
    questions_file_exists = questions_file_path.is_file()
    answers_file_exists = answers_file_path.is_file()

    # Report the status of the files
    if questions_file_exists:
        print(f"Found questions file: {questions_file}")
        print(
            f"Created on: {datetime.fromtimestamp(questions_file_path.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        print(f"Error: Questions file not found: {questions_file}")

    if answers_file_exists:
        print(f"Found answers file: {answers_file}")
        print(f"Created on: {datetime.fromtimestamp(answers_file_path.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        print(f"Error: Answers file not found: {answers_file}")

    # Proceed only if both files exist
    if not (questions_file_exists and answers_file_exists):
        print("Error: Cannot proceed without both files.")
        return None

    # Extract the base filename without the '_dictionary.py' part
    base_filename_questions = questions_file.split('_que_questions_extraction_completed_dictionary.py')[0]
    base_filename_answers = answers_file.split('_ans_correct_answers_dictionary.py')[0]

    # Define the dictionary names
    questions_dict_name = f"{base_filename_questions}_que_questions_extraction_completed_dictionary"
    answers_dict_name = f"{base_filename_answers}_ans_correct_answers_dictionary"

    print(f"Trying to access questions dictionary: {questions_dict_name}")
    print(f"Trying to access answers dictionary: {answers_dict_name}")


    questions_dict_module = load_module(questions_file, folder_path)
    answers_dict_module = load_module(answers_file, folder_path)

    # Access the dictionaries
    questions_dict = getattr(questions_dict_module, questions_dict_name, None)
    answers_dict = getattr(answers_dict_module, answers_dict_name, None)

    if questions_dict is None:
        print(f"Error: Dictionary not found in questions file: {questions_dict_name}")
    if answers_dict is None:
        print(f"Error: Dictionary not found in answers file: {answers_dict_name}")

    if questions_dict is None or answers_dict is None:
        print("Error: One or both dictionaries not found. Joining not possible.")
        return None

    # Initialize the combined dictionary
    questions_with_correct_answers_dictionary = {}

    # Convert question_number to integer in answers_dict for reliable matching
    answers_dict_by_qn = {int(details.get('question_number', 0)): details for _, details in answers_dict.items()}

    # Iterate through questions_dict to merge information
    for q_id, q_info in questions_dict.items():
        question_number = int(q_info['question_number'])  # Ensure integer comparison
        correct_answer_info = answers_dict_by_qn.get(question_number)

        # If a corresponding answer is found, merge its information
        if correct_answer_info:
            # Ensure the correct field names
            # q_info['correct_answer_choice'] = correct_answer_info['correct_answer_choice']
            q_info['correct_answer_choice'] = correct_answer_info.get('correct_answer_choice', 'N/A')
            # q_info['correct_answer_description'] = correct_answer_info['correct_answer_description']
            q_info['correct_answer_description'] = correct_answer_info.get('correct_answer_description', 'N/A')
            # merged_info = {**q_info, **correct_answer_info}  # Merge two dictionaries
            # questions_with_correct_answers_dictionary[q_id] = merged_info
        else:
            # If no corresponding answer is found, just use the question info
            # questions_with_correct_answers_dictionary[q_id] = q_info
            # If no corresponding answer is found, ensure the fields are still added
            q_info['correct_answer_choice'] = 'N/A'
            q_info['correct_answer_description'] = 'N/A'

        # Debugging: Check if the correct answer fields are being set
        print(f"Question Number: {question_number}")
        print(f"Assigned Correct Answer Choice: {q_info['correct_answer_choice']}")
        print(f"Assigned Correct Answer Description: {q_info['correct_answer_description']}")
        questions_with_correct_answers_dictionary[q_id] = q_info
    # Now questions_with_correct_answers_dictionary contains information from both dictionaries

    input(f'Please check the above merging, Press Enter to continue......')
    # Extract the base filename without the 'que_' portion
    # base_filename_questions_modified = "_".join(base_filename_questions.split("_")[0:5])
    base_filename_questions_modified = "_".join(questions_dict_name.split("_")[0:8])
    # input(f"The questions_dict_name is: {questions_dict_name}, Press Enter to continue....")
    # input(f"The base_filename_questions_modified is: {base_filename_questions_modified}, Press Enter to continue....")

    # Write the combined dictionary to a file
    new_dict_name = f'{base_filename_questions_modified}_questions_with_correct_answers_dictionary'
    file_path = folder_path / f'{new_dict_name}.py'

    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(f'{new_dict_name} = ' + repr(questions_with_correct_answers_dictionary))

    print(f"The file '{new_dict_name}.py' has been created in the directory {folder_path}")
    print(f"{new_dict_name}")

    # Load the dictionary using load_dictionary and print it
    loaded_dict = load_dictionary(new_dict_name, folder_path)
    print("Loaded dictionary:", loaded_dict)
    input(f"Please see the dictionary {new_dict_name} soon after it's creation using load_dictionary function, Press Enter to continue.....")

    # Load the dictionary using load_module and print it
    loaded_module = load_module(f"{new_dict_name}.py", folder_path)
    loaded_dict = getattr(loaded_module, new_dict_name, {})
    print("Loaded dictionary:", loaded_dict)
    input(f"Please see the dictionary {new_dict_name} soon after it's creation using load_module function, Press Enter to continue.....")

    # create_excel_for_dict(data_dict, output_dir, filename_prefix)
    create_excel_for_dict(questions_with_correct_answers_dictionary, folder_path, new_dict_name)

    return questions_with_correct_answers_dictionary

def create_hyperlink(base_filename, dir_path):
    file_name = f"{base_filename}_ques_cleaned_temp_combined.txt"
    folder_name = f"Split Pages {base_filename.split('_')[-1]} Questions"
    folder_path = dir_path.parent / folder_name
    file_path = folder_path / file_name
    return file_path.as_uri()

def analyse_extraction(questions_with_correct_answers_dictionary, target, filename_prefix, dir_path):
    # Sort and get all existing question numbers
    all_question_numbers = sorted([int(question.get('question_number')) for question in questions_with_correct_answers_dictionary.values()])

    # Find missing question numbers
    missing_question_numbers = [num for num in range(1, target + 1) if num not in all_question_numbers]
    # questions_without_correct_answer = [question['question_number'] for question in questions_with_correct_answers_dictionary.values() if 'correct_answer_choice' not in question]

    # Questions without correct answer (check for 'N/A' as well)
    questions_without_correct_answer = [
        question['question_number']
        for question in questions_with_correct_answers_dictionary.values()
        if question.get('correct_answer_choice') in [None, 'N/A']
    ]

    # Print statements for debugging
    print("All question numbers:", all_question_numbers)
    print("Missing question numbers:", missing_question_numbers)
    print("Questions without correct answer:", questions_without_correct_answer)

    valid_question_number_list = []
    invalid_question_number_list = []
    valid_correct_answer_list = []
    invalid_correct_answer_list = []
    valid_both_list = []
    invalid_both_list = []

    # Add a check to identify which entry is causing the issue
    for a_id, a_info in questions_with_correct_answers_dictionary.items():
        question_number = int(a_info['question_number'])
        correct_answer_choice = a_info.get('correct_answer_choice', 'N/A')
        question_number = a_info.get('question_number', 'N/A')

        if 0 < question_number < 2024:
            valid_question_number_list.append(a_id)
        else:
            invalid_question_number_list.append(a_id)

        if correct_answer_choice in [None, 'N/A']:
            invalid_correct_answer_list.append(a_id)
            # Debug print
            print(f"Error: Entry with question_number {question_number} does not have 'correct_answer_choice'")
        elif correct_answer_choice in ['a', 'b', 'c', 'd']:
            # Debug print
            # input(f"For question_number: {question_number}, the correct_answer_choice is: {correct_answer_choice}, Press Enter to continue....")
            valid_correct_answer_list.append(a_id)
            print(f"For question_number: {question_number}, the correct_answer_choice is: {correct_answer_choice}, Press Enter to continue....")

        if 0 < question_number < 2024 and correct_answer_choice in ['a', 'b', 'c', 'd']:
            valid_both_list.append(a_id)


    # Count valid numbers and answer choices
    valid_question_numbers = [num for num in all_question_numbers if num <= target]

    # Count the valid question numbers as the target minus the count of missing question numbers
    # valid_question_number_count = target - len(missing_question_numbers)
    # valid_answer_choice_count = len([question for question in questions_with_correct_answers_dictionary.values() if question.get('correct_answer_choice') in ['a', 'b', 'c', 'd']])
    # valid_answer_choice_count = len([
    #    question for question in questions_with_correct_answers_dictionary.values()
    #    if question.get('correct_answer_choice') in ['a', 'b', 'c', 'd']
    # ])
    valid_question_number_count = len(valid_question_number_list)
    valid_answer_choice_count = len(valid_correct_answer_list)

    input(F"The valid_answer_choice_count is: {valid_answer_choice_count}, Press Enter to continue....")

    # Calculate percentages
    valid_question_number_percentage = round((valid_question_number_count / target) * 100, 2)
    valid_answer_choice_percentage = round((valid_answer_choice_count / target) * 100, 2)
    valid_both_percentage = round((valid_question_number_count / target) * 100, 2)

    # Count questions with and without valid years
    current_year = datetime.now().year
    # valid_year_questions = [question for question in questions_with_correct_answers_dictionary.values() if question.get('exam_year') and question.get('exam_year').isdigit() and 1900 <= int(question.get('exam_year')) <= current_year]
    valid_year_questions = [question for question in questions_with_correct_answers_dictionary.values() if question.get('exam_year') and isinstance(question.get('exam_year'), int) and 1900 <= int(question.get('exam_year')) <= current_year]
    invalid_year_questions = [question for question in questions_with_correct_answers_dictionary.values() if question.get('exam_year') == 'YYY' or not (isinstance(question.get('exam_year'), int) and 1900 <= int(question.get('exam_year')) <= current_year)]

    # invalid_year_questions = [question for question in questions_with_correct_answers_dictionary.values() if question.get('exam_year') == 'YYY']

    valid_year_count = len(valid_year_questions)
    invalid_year_count = len(invalid_year_questions)
    valid_year_percentage = round((valid_year_count / target) * 100, 2)
    questions_without_valid_year = [question['question_number'] for question in invalid_year_questions]

    # Current time for the performance summary
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # performance_message = f"Extraction Performance for overall_extraction_performance at {current_time}"

    # Prepare the new row of data for the DataFrame
    new_data = {
        'Sl No.': 1,  # Assuming this is the first row
        'Performance Date and Time': datetime.now().strftime('%d/%m/%Y %I:%M %p'),
        'Target': target,
        'Valid Question Numbers': valid_question_number_count,
        '%age Valid Question Numbers': valid_question_number_percentage,
        'Valid Correct Answers': valid_answer_choice_count,
        '%age Valid Correct Answers': valid_answer_choice_percentage,
        'Both Valid %': valid_both_percentage,
        'Missing Question Numbers': ', '.join(map(str, missing_question_numbers)),
        'Questions Without Correct Answer': ', '.join(map(str, questions_without_correct_answer)),
        'Number of Questions with Valid Year': valid_year_count,
        'Number of Questions without Valid Year': invalid_year_count,
        '%age of Questions with Valid Year': valid_year_percentage,
        'Question Numbers without Valid Year': ', '.join(map(str, questions_without_valid_year)),
        'Questions Text File Hyperlink': f'{dir_path}/{filename_prefix}_que_extracted_questions.txt',# Updated to a valid path
        'Answers Text File Hyperlink': f'{dir_path}/{filename_prefix}_ans_cleaned.txt'
    }

    # Define the DataFrame and save path
    revised_filename_prefix = "_".join(filename_prefix.split("_")[0:8])
    excel_file_path = dir_path / f"{filename_prefix}_overall_extraction_performance.xlsx"
    if excel_file_path.is_file():
        df = pd.read_excel(excel_file_path)
        new_data['Sl No.'] = len(df) + 1
        # sequence used is to ensure that the new data is written in rows added before the existing rows
        df = pd.concat([pd.DataFrame([new_data]), df], ignore_index=True)
    else:
        df = pd.DataFrame([new_data])

    # Save the DataFrame back to the Excel file
    df.to_excel(excel_file_path, index=False, engine='openpyxl')

    # Fix the hyperlink appearance using openpyxl
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Fix hyperlinks for "Questions Text File Hyperlink"
    questions_hyperlink_col = df.columns.get_loc('Questions Text File Hyperlink') + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, questions_hyperlink_col)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color='0000FF', underline='single')
            cell.value = "Click Here"  # Or keep the original URL text

    # Fix hyperlinks for "Answers Text File Hyperlink"
    answers_hyperlink_col = df.columns.get_loc('Answers Text File Hyperlink') + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, answers_hyperlink_col)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color='0000FF', underline='single')
            cell.value = "Click Here"  # Or keep the original URL text
    wb.save(excel_file_path)
    print(f"{excel_file_path.name} Excel file updated in {dir_path}.")


if __name__ == "__main__":
    module_name = os.path.splitext(os.path.basename(__file__))[0]
    print(f"Now you are in {module_name}")
    # input(f"Press Enter to continue in {module_name}")
    parser = argparse.ArgumentParser(description="Join extracted questions with correct answers.")
    parser.add_argument("subfolder", type=str, help="Folder path where the dictionary files are located.")
    parser.add_argument("questions_file", type=str, help="Questions filename.")
    parser.add_argument("answers_file", type=str, help="Answers filename.")
    parser.add_argument("file_name_prefix", type=str, help="file_name_prefix.")
    parser.add_argument("questions_extraction_target", type=float, help="questions_extraction_target.")
    args = parser.parse_args()

    # questions_file = Path(args.questions_file)
    # answers_file = Path(args.answers_file)
    subfolder = Path(str(args.subfolder))
    file_name_prefix = args.file_name_prefix
    target = int(args.questions_extraction_target)

    # input(f"The file_name_prefix passed to {module_name} is: {file_name_prefix}, Press Enter to continue....")
    # input(f"The questions_extraction_target passed to {module_name} is: {target}, Press Enter to continue....")

    """questions_file_name = "disha_26_2020_upsc_pre_gs1_history_ancient_que_questions_extraction_completed_dictionary"
    subfolder = Path('C:/Users/HP/Desktop/Question Bank Trials/batch/Batch_disha_26_2020/Area_history/history_ancient')
    file_name_prefix = "disha_26_2020_upsc_pre_gs1_history_ancient_que"""

    stripped_file_name_prefix = '_'.join(file_name_prefix.split('_')[:-2])
    # input(f"In the module {module_name} stripped_file_name_prefix is: {stripped_file_name_prefix}, Press Enter to continue....")

    questions_file = f"{stripped_file_name_prefix}_que_questions_extraction_completed_dictionary.py"
    answers_file = f"{stripped_file_name_prefix}_ans_correct_answers_dictionary.py"

    print(f"Constructed questions file: {questions_file}.py")
    print(f"Constructed answers file: {answers_file}.py")

    # input("Press Enter to proceed with join questions and answers .....")

    # Initialize questions_with_correct_answers_dictionary
    questions_with_correct_answers_dictionary = None

    questions_with_correct_answers_dictionary = join_questions_with_answers(subfolder, questions_file, answers_file)

    # dict_choice = input(f"Do you want to use the existing dictionary {stripped_file_name_prefix}_questions_with_correct_answers_dictionary.py)? (y/n): ").strip().lower()
    if questions_with_correct_answers_dictionary is not None:
        # Proceed with analysis or further processing
        print("Questions with correct answers have been successfully combined.")
    else:
        # Exit if joining was unsuccessful
        print("Exiting...")
        exit()

    # Prompt for user input to start analysis
    # start_analysis = input("Do you want to start the analysis? (y/n): ").strip().lower()
    # Ninput("Press Enter to continue to start the analysis ....")
    # start_analysis = 'y'
    # Call the analyse_extraction function
    # target = int(input("Enter the target: "))
    # input(f"For analysis, the target is: {target}, Press Enter to continue.....")
    dir_path = Path(str(subfolder))
    # action_choice = input("Choose an action: 1. Join questions and answers 2. Move to analysis (Enter 1 or 2): ").strip()
    # dict_choice = input(f"Do you want to use the existing dictionary {stripped_file_name_prefix}_questions_with_correct_answers_dictionary.py)? (y/n): ").strip().lower()
    # input(f"Press Enter to continue to use the existing dictionary {stripped_file_name_prefix}_questions_with_correct_answers_dictionary.py.....")
    dict_choice = 'y'
    dict_to_be_loaded = f"{stripped_file_name_prefix}_questions_with_correct_answers_dictionary"
    questions_with_correct_answers_dictionary = load_module(dict_to_be_loaded + '.py', subfolder).__dict__[dict_to_be_loaded]

    # Debug print
    input("Please see situation just before invoking analyse_extraction and printing of dictionary to start. Press Enter to continue....")
    # print(questions_with_correct_answers_dictionary)
    analyse_extraction(questions_with_correct_answers_dictionary, int(target), stripped_file_name_prefix, dir_path)
