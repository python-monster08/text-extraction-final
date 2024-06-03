# this module creates an excel for a dictionary based on headers/ columns and their sequence defined by the user
# it uses xlsxwriter instead of openpyxl as the latter was not able to produce drop down options.
# other features of this module are:
# 1. The load_dictionary function returns None if the dictionary file does not exist. This allows
# distinguishing between a non-existent dictionary and an empty dictionary.
# 2.The create_excel_with_message function creates an Excel file with a single message in the first cell
# if the dictionary does not exist or is empty.
# 3. The use of get method ensures that missing fields in the dictionary are handled gracefully by
# providing a default empty string value.
# 4. Before writing a value to a cell, if the value is a list then use is done of isinstance(value, list)
# and if it is a list, convert it to a JSON string using json.dumps(value).
# 5. for "source_file_hyperlink", the write_url method is used to write the hyperlink to the cell.
# 6. to define the sequence only once, a list of headers is used and headers are written to the
# worksheet in a loop with their respective column numbers, without the need for defining the
# column in each worksheet.write() command
# 7. it also has the function create_excel_for_dict which is imported from the module fresh_integrated_identify_all_types
# and renamed as create_excel_from_integrated to have the option of creating an excel from a dictionary with
# sequence of headers and values the way they exist in that dictionary



import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import importlib.util
from openpyxl.worksheet.datavalidation import DataValidation
import xlsxwriter
import json
import re
from datetime import datetime
from openpyxl import Workbook
import sys
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
import argparse
from fresh_integrated_identify_all_types import create_excel_for_dict as create_excel_from_integrated

# Initialize the data list at the beginning of the script or before any conditional blocks where it's used
data = []


def load_dictionary(file_name, base_dir_path):
    """ Load a dictionary from a Python file within the specified directory. """
    dict_path = base_dir_path / f"{file_name}.py"
    if not dict_path.exists():
        print(f"File {dict_path} not found.")
        return {}  # Return an empty dictionary instead of None

    spec = importlib.util.spec_from_file_location(file_name, dict_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    # Ensure the dictionary name does not include the '.py' extension
    dict_name = file_name if not file_name.endswith('.py') else file_name[:-3]

    # Attempt to get the dictionary
    loaded_dict = getattr(module, dict_name, {})

    # Check if the dictionary is empty and print a message
    if not loaded_dict:
        print(f"Loaded dictionary named '{dict_name}' is empty.")
    else:
        print(f"Loaded dictionary named '{dict_name}' successfully.")

    return loaded_dict

def create_excel_with_message(output_excel_file_path, message):
    workbook = xlsxwriter.Workbook(output_excel_file_path)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, message)
    workbook.close()
    print(f"Excel file '{output_excel_file_path.name}' created with message: {message}")

def create_user_defined_format_excel_for_dict(loaded_dict, output_excel_file_path):
    # Create a new Excel workbook file and add a worksheet
    workbook = xlsxwriter.Workbook(output_excel_file_path)
    worksheet = workbook.add_worksheet()

    # Write the headers
    headers = [
        "question_id",
        "question_number",
        "batch_name",
        "exam_name",
        "exam_stage",
        "type_of_question",
        "subject_name",
        "area_name",
        "part_name",
        "exam_year",
        "question_number",
        "question",
        "question_part",
        "answer",
        "answer_part",
        "question_type",
        "question_sub_type",
        "question_part_first_part",
        "list_1_present",
        "list_2_present",
        "list_1_name",
        "list_2_name",
        "question_part_first_part",
        "list_1_row1",
        "list_2_row1",
        "list_1_row2",
        "list_2_row2",
        "list_1_row3",
        "list_2_row3",
        "list_1_row4",
        "list_2_row4",
        "list_1_row5",
        "list_2_row5",
        "list_1_row6",
        "list_2_row6",
        "list_1_row7",
        "list_2_row7",
        "list_1_row8",
        "list_2_row8",
        "list_1_row9",
        "list_2_row9",
        "question_part_third_part",
        "answer_option_a",
        "answer_option_b",
        "answer_option_c",
        "answer_option_d",
        "correct_answer_choice",
        "correct_answer_description",
        "marks",
        "negative_marks",
        "Free/ source_file_hyperlink",
        "test_series",
        "reference",
        "question_part_list1_max_row",
        "question_complete",
        "list_1_entries",
        "list_2_entries",
        "second_stage_text_to_process",
        "third_stage_text_to_process",
        "fourth_stage_text_to_process",
        "fifth_stage_text_to_process",
        "sixth_stage_text_to_process",
        "seventh_stage_text_to_process",
        "eighth_stage_text_to_process",
        "ninth_stage_text_to_process",
        "tenth_stage_text_to_process",
        "eleventh_stage_text_to_process",
        "twelfth_stage_text_to_process",
        "thirteenth_stage_text_to_process",
        "fourteenth_stage_text_to_process",
        "fifteenth_stage_text_to_process",
        "sixteenth_stage_text_to_process",
        "part_before_list_2",
        "part_before_A",
        "dropdown_options",
        "questions_source_file_hyperlink",
        "answers_source_file_hyperlink",
    ]

    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write the data from the dictionary
    row = 1

    for question_id, question_info in loaded_dict.items():
        # print(f"Writing data for Question ID: {question_id}")  # Debugging print statement
        for col, header in enumerate(headers):
            value = question_info.get(header, "")
            if isinstance(value, list):
                value = json.dumps(value)
            if header == "questions_source_file_hyperlink":
                hyperlink = question_info.get(header, "")
                if hyperlink:
                    worksheet.write_url(row, col, hyperlink, string="Click Here")
                else:
                    worksheet.write(row, col, "")
            elif header == "answers_source_file_hyperlink":
                hyperlink = question_info.get(header, "")
                if hyperlink:
                    worksheet.write_url(row, col, hyperlink, string="Click Here")
                else:
                    worksheet.write(row, col, "")
            else:
                # worksheet.write(row, col, question_info.get(header, ""))
                # to handle lists and json.dumps(value)
                worksheet.write(row, col, value)
        row += 1

    # Apply data validation for dropdowns
    dropdown_options = ['simple_question', 'list_type1', 'list_type2', 'r_and_a']

    worksheet.data_validation(f'BW2:BW{row + 1}', {  # Adjust the range to include the last row
        'validate': 'list',
        'source': dropdown_options,
        'input_message': 'Please select an option',
        'error_message': 'Invalid option',
        'dropdown': True
    })

    # Save and close the workbook
    workbook.close()
    print(f"Excel file '{output_excel_file_path.name}' with dropdown list in column F has been created in {output_excel_file_path}")
    exit()


def main():
    module_name = os.path.splitext(os.path.basename(__file__))[0]
    input(f"Now you are in {module_name}, Press Enter to continue.......")
    # input(f"Press Enter to continue in {module_name}")

    # Get arguments passed from the main script
    parser = argparse.ArgumentParser(description="Create Excel from dictionary.")
    parser.add_argument("subfolder", type=str, help="Folder where the dictionary module is located")
    parser.add_argument("created_dict_name", type=str, help="Name of the created dictionary")
    args = parser.parse_args()

    base_dir_path = Path(args.subfolder)
    base_dir_path.mkdir(parents=True, exist_ok=True)
    suggested_dict_name = args.created_dict_name

    # Check if a suggested dictionary name is provided
    if suggested_dict_name:
        print(f"Suggested dictionary: {suggested_dict_name}, Press Enter to proceed with it ......")
        # user_choice = input(f"Proceed with suggested dictionary '{suggested_dict_name}, (y/n)")
        # input(f"Press Enter to proceed with suggested ......")
        user_choice = "y"
        selected_file = suggested_dict_name

        # if user_choice.lower() == "y":
        #     selected_file = suggested_dict_name
        # else:
        #     print("Exiting.")
        #    exit()

    # input(f"Processing Option: {user_choice}, Press Enter to Continue")

    # Load the dictionary
    loaded_dict = load_dictionary(selected_file, base_dir_path)

    # Define the output Excel file path which is the path to the excel
    output_excel_file_path = base_dir_path / f"{selected_file}.xlsx"
    # handling edge case of a non-existent dictionary or a blank dictionary
    if loaded_dict is None:
        create_excel_with_message(output_excel_file_path, "Dictionary doesn't exist")
        return
    elif not loaded_dict:
        create_excel_with_message(output_excel_file_path, "Empty dictionary")
        return

    # Add this to check the dictionary content before writing to the Excel file
    # for question_id, question_info in loaded_dict.items():
    #     print(f"Question ID: {question_id}, Question Sub Type: {question_info.get('question_sub_type', '')}")

    # Ask the user which format to create the Excel file in
    # user_format_choice = input("Create Excel in (1) User-Defined Format or (2) Default Dictionary Format? ")
    print("Creating Excel in User-Defined Format")
    user_format_choice = "1"
    if user_format_choice == "1":
        create_user_defined_format_excel_for_dict(loaded_dict, output_excel_file_path)
    elif user_format_choice == "2":
        create_excel_from_integrated(loaded_dict, base_dir_path, selected_file)
    else:
        print("Invalid choice, exiting.")
        exit()


if __name__ == "__main__":
    main()
